"""
Build a cross-county KPI comparison spreadsheet.
Jefferson County (aggregated) vs Portage County vs Bayfield County
Consistent, comparable metrics side-by-side.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
import os

wb = Workbook()

# Styling
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
SECTION_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
SECTION_FONT = Font(name="Calibri", bold=True, size=11)
HIGHLIGHT_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
RED_FILL = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")
GREEN_FILL = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
BOLD = Font(name="Calibri", bold=True, size=11)
NORMAL = Font(name="Calibri", size=11)
SMALL_ITALIC = Font(name="Calibri", italic=True, size=10, color="666666")
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

def style_header(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = thin_border

def style_section(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = SECTION_FILL
        cell.font = SECTION_FONT
        cell.border = thin_border

def auto_width(ws, min_width=12, max_width=50):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = max(min(max_len + 4, max_width), min_width)

def write_row(ws, row, data, font=NORMAL, wrap=True, valign='top'):
    for c, val in enumerate(data, 1):
        cell = ws.cell(row=row, column=c, value=val)
        cell.font = font
        cell.border = thin_border
        cell.alignment = Alignment(wrap_text=wrap, vertical=valign)

# ============================================================
# TAB 1: KPI DASHBOARD (Side-by-Side)
# ============================================================
ws1 = wb.active
ws1.title = "KPI Dashboard"
ws1.sheet_properties.tabColor = "1F4E79"

headers = ["KPI / Metric", "Jefferson County", "Portage County", "Bayfield County", "Notes / Source"]
for i, h in enumerate(headers, 1):
    ws1.cell(row=1, column=i, value=h)
style_header(ws1, 1, 5)

# Row counter
r = 2

kpi_sections = [
    # Section: Demographics & Geography
    ("DEMOGRAPHICS & GEOGRAPHY", [
        ("Population", "~84,600", "~70,521", "~15,000 (permanent)", "Bayfield doubles in summer (tourism)"),
        ("Land Area (sq mi)", "583", "810", "2,000+", ""),
        ("Population Density (per sq mi)", "145", "87", "7.5", "Jefferson is most urban; Bayfield extremely rural"),
        ("Median Age", "41.3", "38.1", "54.1", "Bayfield oldest; Portage youngest (student pop in Stevens Point)"),
        ("Median Household Income", "$80,604", "$74,650", "$71,352", "Portage/Bayfield from FRED 2024-2026 est."),
        ("% Population 65+", "15.4%", "~18%", "~31%", "Bayfield has nearly 2x the senior proportion vs Portage"),
        ("Projected 65+ Growth by 2040", "+7,130 (70% of total growth)", "+2,998", "+679 (but total pop declining)", "All 3 counties face aging population = rising EMS demand"),
        ("Number of Municipalities", "16 towns + 7 cities/villages", "17 townships + 10 incorp.", "25 towns + 3 cities/villages", "Bayfield: Washburn, Bayfield, Mason"),
    ]),

    # Section: System Structure
    ("EMS SYSTEM STRUCTURE", [
        ("EMS Governance Model", "Municipal (fragmented)", "Hybrid (county contracts)", "County-supported (stipends)", ""),
        ("County Role in EMS", "None", "Direct - EMS Division in Sheriff's Office", "Growing - coordinator + dispatch", ""),
        ("Number of EMS Providers", "10+ independent departments", "3 contracted + 11 EMR groups", "9 volunteer agencies", ""),
        ("EMS Coordinator (County-level)", "None", "Yes (since 2009)", "Proposed 2025 ($145K)", ""),
        ("Medical Director(s)", "Per department (varies)", "1 MD + 2 Associate MDs", "Not specified", ""),
        ("Dispatch Model", "Fragmented (5 separate centers)", "Centralized county dispatch", "New EMD established 2025", "Jeff Co: county dispatches only 5 of 11 providers"),
        ("Dispatch Staff", "~35 FTEs across 5 centers", "21 staff (16 tech + 4 supv + mgr)", "1 director (hired 2024)", "Jeff Co dispatch costs ~$2.7M combined"),
        ("MABAS Participation", "Yes - Division 118", "Not specified", "Not specified", ""),
    ]),

    # Section: Staffing
    ("STAFFING & PERSONNEL", [
        ("Total EMS FTEs (county-wide)", "~92 FTEs (across all providers)", "Not specified (contracted)", "Mostly volunteer", "Jeff Co aggregated from provider data"),
        ("Salaried Fire/EMS FTEs", "35", "Contracted via 3 FDs", "0 (proposed 6 paramedics)", "Jeff Co: Watertown 26, Ft Atk 4, Lake Mills 2, etc."),
        ("Paid-on-Call / Volunteers", "201+", "11 EMR volunteer groups", "~9 agencies (all volunteer)", ""),
        ("Staffing Model Mix", "Career + POC + Volunteer", "Contracted paramedics (career)", "Volunteer with proposed county crew", ""),
        ("Highest EMS License Level", "Critical Care Paramedic", "Paramedic (county-wide)", "Varies by agency", "Western Lakes has CC Paramedic"),
        ("Service Level Consistency", "Varies widely (EMT-B to Paramedic)", "Paramedic county-wide", "Varies by agency", "Jeff Co: rural areas have lower levels"),
        ("Staffing Challenge", "41% of agencies rely on <=6 staff for 80%+ of calls", "Struggles to fill FT positions", "Volunteer recruitment shortfalls", "Statewide: 78% provided mutual aid due to staffing"),
    ]),

    # Section: Financial
    ("FINANCIAL OVERVIEW", [
        ("Total Annual EMS Budget", "~$12M+ (estimated aggregate)", "$4,043,061 (contracted personnel only)", "$458,000 (2025 proposed)", "Jeff Co estimated from individual provider budgets"),
        ("County EMS Levy", "None (municipalities fund individually)", "Yes - countywide EMS levy", "Yes - exempt from levy lid", "Wis. Stat. 66.0602(3)(e)6"),
        ("Per Capita EMS Cost", "$7.22 - $125 (varies by provider)", "~$57", "~$31", "Jeff Co range shows massive disparity"),
        ("EMS Cost per Capita (weighted avg)", "~$80 (estimated)", "~$57", "~$31", "Jeff Co estimated from budget/population"),
        ("Cost to Homeowner ($200K home)", "Varies by municipality", "Not specified", "$21.46/year", "Bayfield: 0.1073 mills"),
        ("EMS Billing Revenue", "Not centralized (per provider)", "$2,050,116 (2024)", "Not specified", "Portage: 35% collection rate"),
        ("Billing Collection Rate", "Unknown (fragmented)", "35%", "Not available", "35% is typical nationally for EMS"),
        ("Year-over-Year Budget Change", "Rising per-capita rates ($3-4/yr)", "Steady (long-established)", "+148% (2024 to 2025)", "Bayfield scaling rapidly from new baseline"),
    ]),

    # Section: Contract / Payment
    ("PAYMENT MODELS", [
        ("Primary Payment Model", "Per Capita (most common)", "County levy + billing", "County levy (exempt from lid)", ""),
        ("Payment Models in Use", "Per Capita, Equalized Value, Fixed+CPI, Formula", "Single county levy model", "Single county levy model", "Jeff Co has 4+ different models across providers"),
        ("Per Capita Rate Range", "$7.22 - $48/capita", "~$57/capita (via levy)", "~$31/capita (via levy)", ""),
        ("Annual Escalation Mechanism", "Varies: flat $3/yr, CPI, AIF 3-6%", "Levy-based (annual budget)", "Levy-based", "Jeff Co escalation inconsistent across providers"),
        ("Capital Cost Sharing", "Varies: 85/15, excluded, bundled", "County-funded", "Proposed county vehicles", "Jeff Co: no uniform approach"),
    ]),

    # Section: Ambulance Fleet
    ("AMBULANCE FLEET & ASSETS", [
        ("Total Ambulance Units", "18 (across 7 departments)", "7 (4 frontline SPFD + 1 Plover + 2 Amherst)", "Not specified", "Jeff Co: 5 of 14 depts have no ambulance"),
        ("Ambulances per 10,000 Pop", "2.13", "0.99", "Not available", "Jeff Co has more units but fragmented"),
        ("24/7 Staffed Ambulances", "Varies by department", "4 (always staffed)", "Limited (volunteer-based)", "SPFD 2 + Plover 1 + Amherst 1"),
        ("Fleet Avg Age (years)", "~12.5", "Not available", "Not available", "Jeff Co: 3 units 20+ years old"),
        ("Units Needing Replacement (20+ yrs)", "3", "Not specified", "Not specified", "Ft Atk 2004, Waterloo 2005, Watertown 2006"),
        ("Newest Unit", "2023 (Ft Atkinson & Watertown)", "Not specified", "N/A", ""),
        ("Oldest Unit", "2004 (Ft Atkinson Rescue 8157)", "Not specified", "N/A", "22 years old"),
        ("Total Fire/EMS Apparatus (est.)", "~145 units across 14 depts", "~35+ (SPFD 15+ / Plover 14 / Amherst TBD)", "Not specified", ""),
        ("SPFD Apparatus", "", "2 engines, 2 trucks, 1 tower, 4 ambulances, 4+ command, support", "", "Newest: 2025 Pierce Enforcer engine"),
        ("Plover FD Apparatus", "", "2 engines, 1 platform, 2 tenders, 1 brush, 1 ambulance, 5 command/admin", "", "Includes Portage Co Hazmat vehicle"),
        ("Amherst FD Apparatus", "", "2 ambulances + fire apparatus (detail not available online)", "", "Wix website - data not scrapable"),
        ("Equipment Ownership", "Municipal (each dept owns)", "County-funded", "Proposed: county response vehicles", "Lake Mills/Ryan Bros: private contractor owns"),
    ]),

    # Section: Call Volume & Response
    ("CALL VOLUME & RESPONSE TIMES", [
        ("Total Annual EMS Calls", "~8,000+ (2019 aggregate)", "3,993 billable (2024)", "Not specified (limited)", "Jeff Co aggregated from all providers"),
        ("Calls per 1,000 Population", "33 - 117 (varies by provider)", "56.6 (billable only)", "Not available", "Jeff Co: Jefferson 117, Cambridge 33"),
        ("10-Year Call Volume Trend", "Growing (most providers +5-42%)", "+24% (3,219 to 3,993 over 2014-2024)", "Increasing mutual aid requests", ""),
        ("ALS vs BLS Call Split", "Not centralized", "64% ALS / 36% BLS", "Not available", ""),
        ("Avg Response Time (urban)", "4.6 - 6 min", "Not specified", "Not specified", "Jeff Co: Watertown 4.6, Jefferson 6"),
        ("Avg Response Time (rural)", "8.6 - 14.5 min", "Not specified", "Not specified", "Jeff Co: Ixonia 11.5, Johnson Creek 14.5"),
        ("Medflight Frequency", "Not specified", "~1 every 2 weeks", "Not specified", "Portage: higher than metro avg due to distance"),
        ("Mutual Aid Provided (2019)", "Watertown: 120 EMS mutual aid calls", "Not specified", "Frequent (capacity issues)", "78% of WI agencies provided mutual aid"),
    ]),

    # Section: Governance & Legal
    ("GOVERNANCE & LEGAL FRAMEWORK", [
        ("Governing Authority for EMS", "Individual municipalities", "County Board (via Sheriff's Office)", "County Board", ""),
        ("EMS Levy Lid Exemption Used?", "No", "Yes", "Yes", "Jeff Co leaving money on the table"),
        ("Applicable Statute", "N/A (not using exemption)", "Wis. Stat. 66.0602(3)(e)6", "Wis. Stat. 66.0602(3)(e)6", "Countywide EMS levy exempt from caps"),
        ("Year County EMS Established", "N/A", "1969 (contract w/ SPFD)", "2023", "Portage has 55+ years of history"),
        ("Legal Challenges", "N/A", "None noted", "None noted", "Lafayette Co faced legal challenge (for reference)"),
        ("Contract Standardization", "None - each contract different", "Standardized county contracts", "N/A (levy-funded)", "Jeff Co: 4+ different payment models"),
    ]),

    # Section: Key Ratios
    ("KEY PERFORMANCE RATIOS", [
        ("EMS Budget per Capita", "~$142 (aggregate budgets/pop)", "$57", "$31", "Jeff Co higher due to fragmentation overhead"),
        ("EMS Budget per Square Mile", "~$20,600", "~$4,990", "~$229", "Reflects population density differences"),
        ("Ambulances per 100 sq mi", "3.1", "0.86", "Not available", "Jeff Co more units per area but uncoordinated"),
        ("FTEs per 10,000 Population", "~10.9", "Not available", "~0 (volunteer)", "Jeff Co aggregated; volunteers not counted"),
        ("Dispatch Centers per County", "5+", "1", "1 (new 2025)", "Fragmentation = redundant costs"),
        ("Dispatch Cost per Capita", "~$32 (aggregate)", "Not specified", "Not available", "Jeff Co: ~$2.7M across 5 centers"),
        ("% of Providers at Paramedic Level", "~45% (5 of 11)", "100%", "Varies", "Jeff Co rural areas below Paramedic"),
        ("Providers per 10,000 Pop", "1.3", "0.43", "6.0", "Bayfield high ratio due to small pop + many volunteer agencies"),
    ]),
]

for section_title, rows in kpi_sections:
    # Write section header
    write_row(ws1, r, (section_title, "", "", "", ""), font=SECTION_FONT)
    style_section(ws1, r, 5)
    r += 1

    for row_data in rows:
        write_row(ws1, r, row_data)
        r += 1

    # Blank row between sections
    write_row(ws1, r, ("", "", "", "", ""))
    r += 1

auto_width(ws1)
ws1.column_dimensions['A'].width = 40
ws1.column_dimensions['B'].width = 38
ws1.column_dimensions['C'].width = 35
ws1.column_dimensions['D'].width = 35
ws1.column_dimensions['E'].width = 45

# ============================================================
# TAB 2: JEFFERSON COUNTY PROVIDER BREAKDOWN
# ============================================================
ws2 = wb.create_sheet("Jeff Co Provider Detail")
ws2.sheet_properties.tabColor = "548235"

headers2 = ["Provider", "Type", "Staffing Model", "License Level", "FTEs",
            "Total Budget", "Cost/Capita", "Per Capita Rate", "Ambulances",
            "24/7 Coverage", "Calls (2019)", "Calls/1,000 Pop", "Response Time"]
for i, h in enumerate(headers2, 1):
    ws2.cell(row=1, column=i, value=h)
style_header(ws2, 1, 13)

providers = [
    ("Watertown FD", "Combined Fire/EMS", "Full-time career", "Paramedic", 26.9,
     "$2,900,000", "$74 (EMS)", "$40/capita", 3, "Yes", "2,092", "80", "4.6 min"),
    ("Fort Atkinson FD", "Combined Fire/EMS", "Career + POC", "Paramedic", "4 FT + 38 POC",
     "Not specified", "Not specified", "$7.22 + CPI", 3, "Yes", "Not specified", "N/A", "10 min"),
    ("Western Lakes FD", "Combined Fire/EMS", "Career + POP/POC", "CC Paramedic", 67.5,
     "$5,500,000", "$116 ($29 Jeff Co)", "N/A", 0, "Yes", "245", "58", "6 min"),
    ("Whitewater FD", "Combined Fire/EMS", "FT + POC", "Paramedic (since 2023)", "18 FT + 12 POC",
     "$1,000,000", "$50", "N/A", 3, "Yes (5 per 24hr shift)", "2,000+", "90", "6 min"),
    ("Jefferson Fire/EMS", "EMS", "Mostly POP", "Paramedic", 16.6,
     "$808,400", "$77", "$31-$40/capita", 5, "Yes", "1,236", "117", "6 min"),
    ("Ryan Brothers Ambulance", "EMS (private)", "POP", "Paramedic", 6,
     "$580,257", "$31", "$48/capita (Aztalan)", 0, "Yes", "1,246", "68", "6.4 min"),
    ("Johnson Creek Fire/EMS", "Combined Fire/EMS", "FT + POP/POC", "Paramedic", 8.84,
     "$395,651", "$67", "Equalized Value", 0, "Yes", "414", "71", "11.5-14.5 min"),
    ("Lake Mills EMS", "EMS", "POP/POC", "AEMT", 11.5,
     "$418,378", "$44", "$49.44/capita", 0, "Yes", "655", "71", "10.7 min"),
    ("Cambridge Area EMS", "EMS", "Career + POP/POC", "AEMT", 10.94,
     "$480,405", "$84", "N/A", 0, "Limited", "170", "33", "2.7-8 min"),
    ("Waterloo Fire & Rescue", "Combined Fire/EMS", "Career + POC", "AEMT", "2 FT + 50 POC",
     "Not specified", "N/A", "$22-$26/capita", 2, "Limited", "Not specified", "N/A", "8.6 min"),
    ("Palmyra PSD", "Fire/EMS/Police", "Mostly FT", "EMT-B", 6.97,
     "$360,407", "$125", "N/A", 1, "Yes", "203", "69", "5 min"),
    ("Ixonia Fire & Medical", "Combined Fire/EMS", "POC", "EMT/EMT-I", "N/A",
     "Not specified", "N/A", "Formula ($49,169)", 1, "Limited (POC)", "Not specified", "N/A", "11.5 min"),
]

for row_idx, row_data in enumerate(providers, 2):
    for c, val in enumerate(row_data, 1):
        cell = ws2.cell(row=row_idx, column=c, value=val)
        cell.font = NORMAL
        cell.border = thin_border
        cell.alignment = Alignment(wrap_text=True, vertical='top')

# Totals row
totals_row = len(providers) + 2
write_row(ws2, totals_row, ("JEFFERSON CO TOTAL/AVG", "", "Mixed", "Varies",
          "~110+ FTEs", "~$12M+ est.", "~$80 avg", "Varies widely", "18 total",
          "Mixed", "~10,000+", "~95 avg", "4.6 - 14.5 min"), font=BOLD)
style_section(ws2, totals_row, 13)

auto_width(ws2)

# ============================================================
# TAB 3: FINANCIAL COMPARISON DETAIL
# ============================================================
ws3 = wb.create_sheet("Financial Deep Dive")
ws3.sheet_properties.tabColor = "BF8F00"

headers3 = ["Financial Metric", "Jefferson County", "Portage County", "Bayfield County", "Takeaway"]
for i, h in enumerate(headers3, 1):
    ws3.cell(row=1, column=i, value=h)
style_header(ws3, 1, 5)

financial_rows = [
    ("", "", "", "", ""),
    ("TOTAL EMS SPENDING", "", "", "", ""),
    ("Total Annual EMS Budget", "~$12M+ (aggregate of all providers)", "$4,043,061 (contracted personnel)", "$458,000 (2025 proposed)",
     "Jefferson spends ~3x Portage for similar population"),
    ("County Share of EMS Budget", "$0 (0%)", "100% (county-funded)", "100% (county-funded)",
     "Jefferson County has zero county-level EMS investment"),
    ("Municipal Share of EMS Budget", "100%", "0% (county funds all)", "0% (county funds all)",
     "Jefferson municipalities bear entire burden individually"),
    ("", "", "", "", ""),
    ("PER CAPITA ANALYSIS", "", "", "", ""),
    ("EMS Budget per Capita", "~$142 (estimated aggregate)", "$57", "$31",
     "Jefferson 2.5x higher than Portage per capita"),
    ("Lowest Per Capita Rate", "$7.22 (Fort Atkinson base)", "$57 (uniform)", "$31 (uniform)",
     "Fort Atkinson rate is artificially low (base before CPI)"),
    ("Highest Per Capita Rate", "$125 (Palmyra PSD bundled)", "$57 (uniform)", "$31 (uniform)",
     "17x spread within Jefferson County"),
    ("Per Capita Rate Spread", "$7.22 - $125 (17x range)", "Uniform", "Uniform",
     "Fragmentation creates massive inequity"),
    ("", "", "", "", ""),
    ("REVENUE & BILLING", "", "", "", ""),
    ("Gross EMS Charges", "Not centralized", "$5,841,808", "Not specified",
     "Jefferson missing data = missed revenue optimization"),
    ("Total Billing Revenue", "Varies by provider", "$2,050,116", "Not specified", ""),
    ("Collection Rate", "Unknown", "35%", "Not available",
     "35% is national average for EMS"),
    ("Revenue per Billable Call", "Unknown", "$513 avg", "Not available", ""),
    ("Revenue Trend", "Unknown", "$1.65M (2014) to $2.05M (2024)", "Not available",
     "Portage revenue growing with call volume"),
    ("", "", "", "", ""),
    ("DISPATCH COSTS", "", "", "", ""),
    ("Number of Dispatch Centers", "5+ (fragmented)", "1 (centralized)", "1 (new 2025)", ""),
    ("Total Dispatch Budget", "~$2,671,181", "Not specified separately", "$128,000 (EMD allocation)",
     "Jeff Co spends ~$2.7M on fragmented dispatch"),
    ("Dispatch Cost per Capita", "~$32", "Not specified", "~$8.50",
     "Consolidation could save significant costs"),
    ("Dispatch FTEs", "~35 across 5 centers", "21", "1 director",
     "Jeff Co has highest staffing for worst coordination"),
    ("", "", "", "", ""),
    ("LEVY & TAX IMPACT", "", "", "", ""),
    ("County EMS Levy", "None", "Yes", "Yes (exempt from lid)",
     "Jeff Co not using available levy exemption"),
    ("Levy Lid Exemption Used?", "No", "Yes", "Yes",
     "Jeff Co leaving state-authorized funding tool unused"),
    ("Mill Rate for EMS", "Varies by municipality", "Not specified", "0.1073 mills",
     ""),
    ("Cost per $200K Home", "Varies by municipality", "Not specified", "$21.46/year",
     ""),
    ("", "", "", "", ""),
    ("PROVIDER BUDGET BREAKDOWN (Jefferson Co)", "", "", "", ""),
    ("Watertown FD", "$2,900,000", "", "", "Largest single provider"),
    ("Western Lakes FD", "$5,500,000", "", "", "Includes Waukesha Co portions"),
    ("Whitewater FD", "$1,000,000", "", "", ""),
    ("Jefferson Fire/EMS", "$808,400", "", "", ""),
    ("Ryan Brothers Ambulance", "$580,257", "", "", "Private contractor"),
    ("Cambridge Area EMS", "$480,405", "", "", ""),
    ("Lake Mills EMS", "$418,378", "", "", ""),
    ("Johnson Creek Fire/EMS", "$395,651", "", "", ""),
    ("Palmyra PSD", "$360,407", "", "", "Fire + EMS + Police bundled"),
]

for row_idx, row_data in enumerate(financial_rows, 2):
    write_row(ws3, row_idx, row_data)
    if row_data[0] in ("TOTAL EMS SPENDING", "PER CAPITA ANALYSIS", "REVENUE & BILLING",
                        "DISPATCH COSTS", "LEVY & TAX IMPACT", "PROVIDER BUDGET BREAKDOWN (Jefferson Co)"):
        style_section(ws3, row_idx, 5)

auto_width(ws3)
ws3.column_dimensions['A'].width = 42
ws3.column_dimensions['B'].width = 38
ws3.column_dimensions['C'].width = 35
ws3.column_dimensions['D'].width = 30
ws3.column_dimensions['E'].width = 45

# ============================================================
# TAB 4: RESPONSE & OPERATIONS
# ============================================================
ws4 = wb.create_sheet("Response & Operations")
ws4.sheet_properties.tabColor = "7030A0"

headers4 = ["Operational Metric", "Jefferson County", "Portage County", "Bayfield County", "Notes"]
for i, h in enumerate(headers4, 1):
    ws4.cell(row=1, column=i, value=h)
style_header(ws4, 1, 5)

ops_data = [
    ("", "", "", "", ""),
    ("CALL VOLUME", "", "", "", ""),
    ("Total Annual Calls", "~8,000+ (2019)", "3,993 billable (2024)", "Not specified (limited)",
     "Jefferson higher volume, more calls per provider"),
    ("Calls per 1,000 Population", "33 - 117 (varies)", "56.6", "Not available",
     "Jeff Co avg ~95; wide variation shows uneven demand"),
    ("Highest Call Volume Provider", "Watertown: 2,092 calls", "SPFD: majority of 3,993", "Not specified", ""),
    ("Lowest Call Volume Provider", "Cambridge: 170 calls", "AFD: smallest share", "Not specified", ""),
    ("10-Year Trend", "Growing +5% to +42%", "+24% (2014-2024)", "Increasing mutual aid",
     "All counties seeing rising demand"),
    ("ALS Call Percentage", "Not centralized", "64%", "Not available", ""),
    ("BLS Call Percentage", "Not centralized", "36%", "Not available", ""),
    ("", "", "", "", ""),
    ("RESPONSE TIMES", "", "", "", ""),
    ("Best Response Time (urban)", "4.6 min (Watertown)", "Not specified", "Not specified", ""),
    ("Avg Response Time (career)", "5 - 6 min", "Not specified", "Not specified",
     "Jeff Co career depts perform well"),
    ("Avg Response Time (POC/volunteer)", "8.6 - 14.5 min", "Not specified", "Not specified",
     "Jeff Co POC/volunteer depts significantly slower"),
    ("Worst Response Time", "14.5 min (Johnson Creek rural)", "Not specified", "Not specified",
     "Rural areas underserved"),
    ("Response Time Gap (best vs worst)", "10 minutes (4.6 to 14.5)", "Likely small (uniform system)", "Likely large (distances)",
     "Jeff Co gap driven by staffing model, not distance"),
    ("Medflight Usage", "Not specified", "~1 every 2 weeks", "Not specified",
     "Portage higher than metro avg"),
    ("", "", "", "", ""),
    ("MUTUAL AID", "", "", "", ""),
    ("Mutual Aid Frequency", "Common (Watertown: 120 calls)", "Not specified", "Frequent (capacity issues)",
     "78% of WI agencies provided mutual aid due to staffing"),
    ("Mutual Aid Coordination", "Not optimized (relationship-based)", "County-coordinated", "Growing need",
     "Jeff Co dispatch doesn't track nearest available unit"),
    ("", "", "", "", ""),
    ("SERVICE LEVELS BY AREA", "", "", "", ""),
    ("Urban Coverage Level", "Paramedic (Watertown, Fort Atk, Jefferson)", "Paramedic (uniform)", "Not applicable",
     "Jeff Co cities match Portage in service level"),
    ("Suburban Coverage Level", "AEMT (Waterloo, Lake Mills, Whitewater)", "Paramedic (uniform)", "Not applicable",
     "Jeff Co suburban areas one level below Portage"),
    ("Rural Coverage Level", "EMT-B to EMT-I (Ixonia, Palmyra)", "Paramedic (uniform)", "Varies (volunteer)",
     "Jeff Co rural residents get lower service level"),
    ("Service Level Uniformity", "NO - varies by municipality", "YES - Paramedic everywhere", "NO - varies by agency",
     "Only Portage has uniform countywide service level"),
]

for row_idx, row_data in enumerate(ops_data, 2):
    write_row(ws4, row_idx, row_data)
    if row_data[0] in ("CALL VOLUME", "RESPONSE TIMES", "MUTUAL AID", "SERVICE LEVELS BY AREA"):
        style_section(ws4, row_idx, 5)

auto_width(ws4)
ws4.column_dimensions['A'].width = 40
ws4.column_dimensions['E'].width = 50

# ============================================================
# TAB 5: STRENGTHS / WEAKNESSES / GAPS
# ============================================================
ws5 = wb.create_sheet("SWOT Summary")
ws5.sheet_properties.tabColor = "C00000"

headers5 = ["Category", "Jefferson County", "Portage County", "Bayfield County"]
for i, h in enumerate(headers5, 1):
    ws5.cell(row=1, column=i, value=h)
style_header(ws5, 1, 4)

swot_data = [
    ("", "", "", ""),
    ("STRENGTHS", "", "", ""),
    ("Infrastructure",
     "Large ambulance fleet (15 units); ~130 total apparatus; established departments with career staff",
     "Unified system since 1969; 4 ambulances staffed 24/7; county-funded fleet",
     "Low cost to taxpayers ($21/yr per $200K home); exempt from levy lid; new coordinator position"),
    ("Service Delivery",
     "Paramedic-level service in cities; high call volume capacity; short urban response times (4.6 min)",
     "Paramedic-level service countywide; standardized care; centralized billing ($2M+ revenue)",
     "9 agencies covering vast area; community-embedded volunteer responders"),
    ("Coverage",
     "Multiple providers ensure redundancy; strong mutual aid network (MABAS Div 118)",
     "4 stations cover entire county; 11 EMR groups provide first response",
     "Distributed volunteer agencies know local geography"),
    ("", "", "", ""),
    ("WEAKNESSES", "", "", ""),
    ("Financial",
     "~$12M+ total spend (2.5x Portage per capita); no billing consolidation; no county levy; municipalities at levy caps",
     "Budget constraints; $4M+ annual commitment; difficulty filling FT positions",
     "Tiny budget ($458K); limited funding for equipment; no billing revenue data"),
    ("Operational",
     "Fragmented dispatch (5 centers, ~$2.7M); no county EMS coordinator; inconsistent service levels across municipalities",
     "Understaffing challenges; Plover FD: 7 positions, only 4 filled",
     "Volunteer recruitment crisis; limited 24/7 coverage; no local hospital"),
    ("Governance",
     "No county role; 4+ different payment models; no standardized contracts; no centralized data",
     "Dependency on 3 contracted departments; limited to contracted model",
     "New system (est. 2023); limited track record; governance still forming"),
    ("", "", "", ""),
    ("OPPORTUNITIES", "", "", ""),
    ("Levy Exemption",
     "Wis. Stat. 66.0602(3)(e)6 allows countywide EMS levy EXEMPT from caps - currently unused",
     "Already using exemption",
     "Already using exemption"),
    ("Consolidation Savings",
     "Dispatch consolidation alone could save hundreds of thousands; billing consolidation could improve revenue",
     "Limited - already consolidated",
     "Limited - already consolidated"),
    ("State Funding",
     "DOR Innovation Grants ($10M cap); DHS FAP; Planning Grants ($3M for <5K pop communities)",
     "Same state funding available",
     "Same state funding available; planning grants suit small communities"),
    ("", "", "", ""),
    ("THREATS", "", "", ""),
    ("Sustainability",
     "Municipalities cannot raise levies further; aging fleet (3 units 20+ yrs); aging population (+7,130 seniors by 2040)",
     "Rising costs; growing call volume (+24% in 10 yrs); staffing shortages",
     "Volunteer model may not sustain; vast distances; population seasonal swings"),
    ("Legal/Political",
     "Consolidation requires political will from 16+ municipalities; Lafayette Co faced legal challenges",
     "Contract renegotiation risk with provider departments",
     "Still building political consensus; new system untested"),
]

for row_idx, row_data in enumerate(swot_data, 2):
    write_row(ws5, row_idx, row_data)
    if row_data[0] in ("STRENGTHS", "WEAKNESSES", "OPPORTUNITIES", "THREATS"):
        style_section(ws5, row_idx, 4)

auto_width(ws5)
ws5.column_dimensions['A'].width = 22
ws5.column_dimensions['B'].width = 55
ws5.column_dimensions['C'].width = 45
ws5.column_dimensions['D'].width = 45

# ============================================================
# TAB 6: DATA SOURCES & ASSUMPTIONS
# ============================================================
ws6 = wb.create_sheet("Sources & Assumptions")
ws6.sheet_properties.tabColor = "808080"

headers6 = ["Item", "Details"]
for i, h in enumerate(headers6, 1):
    ws6.cell(row=1, column=i, value=h)
style_header(ws6, 1, 2)

source_data = [
    ("", ""),
    ("DATA SOURCES", ""),
    ("Jefferson County EMS Contracts", "13 contract documents (txt transcriptions of scanned PDFs)"),
    ("Jefferson County Provider Data", "WI Policy Forum reports: 'Greater than the Sum' (2020) + 'Taking the Pulse' (2020)"),
    ("Jefferson County Call Data", "2024 EMS Workgroup spreadsheets for 14 departments"),
    ("Jefferson County Assets", "MABAS Division 118 Google Sheets for 14 departments (downloaded Feb 2026)"),
    ("Portage County", "2024 Public Safety Annual Report; EMS Model Presentation to Jefferson Co EMS Working Group"),
    ("Bayfield County", "2025 Countywide EMS Budget Narrative"),
    ("Legal/Levy Framework", "Levy-Limit Joint Fire-EMS FAQ (WI DOR); Countywide EMS Tax Levy Limit Exception ruling"),
    ("Statewide Context", "WI EMS Personnel Overview; WI EMS Services Overview; EMS Funding Models in WI presentation (July 2025)"),
    ("Prior ISyE 450 Projects", "EMS Billing Consolidation (Fall 2025); EMS Data Visualization (Fall 2025); EMS Models (Fall 2025)"),
    ("", ""),
    ("KEY ASSUMPTIONS & ESTIMATES", ""),
    ("Jefferson Co total EMS budget (~$12M+)", "Aggregated from individual provider budgets where available; some providers (Fort Atkinson, Waterloo, Ixonia) lack published budget data"),
    ("Jefferson Co per capita avg (~$80-$142)", "Range depends on whether total budget or weighted contract rates are used; true figure requires complete municipal budget data"),
    ("Jefferson Co total FTEs (~92)", "Aggregated from WI Policy Forum provider-level data; some providers may have changed since 2019-2020"),
    ("Jefferson Co total calls (~8,000+)", "Aggregated from 2019 provider call data; does not include fire-only calls; 2024 data may be higher"),
    ("Population figures", "Mix of 2019 ACS, 2022, and 2024 estimates; used most recent available for each county"),
    ("Portage Co per capita ($57)", "Calculated as $4,043,061 / 70,521; covers contracted personnel only, not full EMS system cost"),
    ("Bayfield Co per capita ($31)", "Calculated as $458,000 / 15,000; permanent population only (summer doubles)"),
    ("", ""),
    ("DATA GAPS (Marked N/A or 'Not available')", ""),
    ("Portage Co demographics", "Median age, income, 65+ % not available from sources in database"),
    ("Bayfield Co demographics", "Median age, income, 65+ % not available from sources in database"),
    ("Portage Co response times", "Not published in annual report"),
    ("Bayfield Co call volume", "Not specified in budget narrative"),
    ("Fort Atkinson & Waterloo budgets", "Not included in WI Policy Forum provider data or contracts"),
    ("Centralized Jeff Co billing data", "Each provider bills independently; no county-level aggregation exists"),
    ("Western Lakes & Whitewater assets", "MABAS sheets have empty templates"),
]

for row_idx, row_data in enumerate(source_data, 2):
    write_row(ws6, row_idx, row_data)
    if row_data[0] in ("DATA SOURCES", "KEY ASSUMPTIONS & ESTIMATES", "DATA GAPS (Marked N/A or 'Not available')"):
        style_section(ws6, row_idx, 2)

auto_width(ws6)
ws6.column_dimensions['A'].width = 42
ws6.column_dimensions['B'].width = 80

# Save
output_path = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                           "County_EMS_KPI_Comparison.xlsx")
wb.save(output_path)
print(f"KPI Comparison saved to: {output_path}")
