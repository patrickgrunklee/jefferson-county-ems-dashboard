"""
Build the 3-county EMS Financial Comparison spreadsheet and report.
Jefferson County (municipality-by-municipality) vs Portage County vs Bayfield County
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
import os

wb = Workbook()

# Color scheme
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
SECTION_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
SECTION_FONT = Font(name="Calibri", bold=True, size=11)
BOLD = Font(name="Calibri", bold=True, size=11)
NORMAL = Font(name="Calibri", size=11)
MONEY = '#,##0'
MONEY_DEC = '#,##0.00'
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

def style_header(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = thin_border

def style_section(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = SECTION_FILL
        cell.font = SECTION_FONT
        cell.border = thin_border

def auto_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 4, 45)

# ============================================================
# TAB 1: COUNTY OVERVIEW COMPARISON
# ============================================================
ws1 = wb.active
ws1.title = "County Overview"
ws1.sheet_properties.tabColor = "2F5496"

headers = ["Metric", "Jefferson County", "Portage County", "Bayfield County"]
for i, h in enumerate(headers, 1):
    ws1.cell(row=1, column=i, value=h)
style_header(ws1, 1, 4)

overview_data = [
    ("", "", "", ""),
    ("GENERAL INFORMATION", "", "", ""),
    ("Population (approx.)", "~84,000", "~70,521", "~15,000"),
    ("Land Area (sq mi)", "~576", "~823", "~2,000+"),
    ("Number of EMS Providers", "10+ municipal departments", "3 contracted + 11 EMR groups", "9 volunteer agencies"),
    ("EMS System Model", "Municipal (no county role)", "Hybrid (county contracts w/ municipalities)", "County-supported (stipends to providers)"),
    ("County EMS Coordinator", "None", "Yes (in Sheriff's Office since 2009)", "Proposed for 2025 ($145,000)"),
    ("EMS Levy Status", "No countywide EMS levy", "Yes - Countywide EMS levy", "Yes - Exempt from levy lid"),
    ("", "", "", ""),
    ("FINANCIAL OVERVIEW", "", "", ""),
    ("Total Annual EMS Budget", "Not consolidated (municipal budgets)", "$4,043,061 (contracted personnel only, 2025)", "$458,000 (2025 proposed)"),
    ("EMS Levy Amount", "N/A - municipalities fund individually", "Funded via county EMS levy", "$185,400 (2024) → $458,000 (2025)"),
    ("EMS Revenue from Billing", "Varies by provider (not centralized)", "$2,050,116 (2024) on $5,841,808 charges", "Not specified"),
    ("Per Capita Cost to Citizens", "Varies: $7.22 - $48/capita by contract", "~$57/capita ($4M / 70,521)", "~$21.46/yr per $200k home value"),
    ("Cost Model", "Per-capita or Equalized Value", "County levy + billing revenue", "County levy (exempt from lid)"),
    ("", "", "", ""),
    ("STAFFING", "", "", ""),
    ("Staffing Model", "Municipal employees + volunteers", "Contracted paramedics (24/7)", "Mostly volunteer responders"),
    ("24/7 Ambulance Coverage", "Yes - via municipal departments", "4 frontline ambulances 24/7", "Limited - volunteer-based"),
    ("Ambulance Units (frontline)", "15 units across 6 depts (MABAS Div 118)", "7 total: SPFD 4 (Med 2-5) + Plover 1 (Med 1) + Amherst 2", "Not specified"),
    ("Medical Director", "Per department", "1 Medical Director + 2 Associates", "Not specified"),
    ("EMS Service Level", "Paramedic (cities), ALS/BLS (towns)", "Paramedic county-wide", "Varies by agency"),
    ("", "", "", ""),
    ("DISPATCH", "", "", ""),
    ("Dispatch Model", "Mix of municipal and county dispatch", "County dispatch center (24/7)", "Not specified"),
    ("Dispatch Staff", "Varies by municipality", "16 FT technicians + 4 supervisors + manager", "Proposed Emergency Medical Dispatch ($128,000)"),
    ("Calls for Service (2024)", "Data in call data spreadsheets", "53,166 total CAD calls", "Not specified"),
    ("", "", "", ""),
    ("KEY CHALLENGES", "", "", ""),
    ("Primary Challenge", "Fragmented system, no coordination", "Budget constraints, understaffing", "Volunteer recruitment, vast distances"),
    ("Levy Situation", "Each muni at/near levy cap", "County levy covers EMS", "Countywide EMS exempt from levy lid"),
]

for r, row_data in enumerate(overview_data, 2):
    for c, val in enumerate(row_data, 1):
        cell = ws1.cell(row=r, column=c, value=val)
        cell.font = NORMAL
        cell.border = thin_border
        cell.alignment = Alignment(wrap_text=True)
    # Style section headers
    if row_data[0] in ("GENERAL INFORMATION", "FINANCIAL OVERVIEW", "STAFFING", "DISPATCH", "KEY CHALLENGES"):
        style_section(ws1, r, 4)

auto_width(ws1)

# ============================================================
# TAB 2: JEFFERSON COUNTY CONTRACT DETAILS
# ============================================================
ws2 = wb.create_sheet("Jefferson Co Contracts")
ws2.sheet_properties.tabColor = "548235"

headers2 = ["Provider", "Town/Municipality Served", "Services", "Payment Model",
            "Rate/Amount (2024-2025)", "Contract Period", "Capital/Vehicle Cost Sharing", "Per-Call Fees"]
for i, h in enumerate(headers2, 1):
    ws2.cell(row=1, column=i, value=h)
style_header(ws2, 1, 8)

contracts = [
    # City of Jefferson contracts
    ("City of Jefferson", "Town of Jefferson", "Paramedic, Advanced, Basic EMS",
     "Per Capita", "$31/capita (2024) → $34 (2025) → $37 (2026) → $40 (2027)",
     "Jan 2023 - Dec 2027", "85% Towns / 15% City for vehicle replacement", "Yes - $75 hook-up fee per fire/hydrant"),
    ("City of Jefferson", "Town of Farmington", "Paramedic, Advanced, Basic EMS",
     "Per Capita", "$31/capita (2024) → $34 (2025) → $37 (2026) → $40 (2027)",
     "Jan 2023 - Dec 2027", "85% Towns / 15% City for vehicle replacement", "Yes - $75 hook-up fee"),
    ("City of Jefferson", "Town of Hebron", "Paramedic, Advanced, Basic EMS",
     "Per Capita", "$31/capita (2024) → $34 (2025) → $37 (2026) → $40 (2027)",
     "Jan 2023 - Dec 2027", "85% Towns / 15% City for vehicle replacement", "Yes - $75 hook-up fee"),
    ("City of Jefferson", "Town of Oakland", "Paramedic, Advanced, Basic EMS",
     "Per Capita", "$31/capita (2024) → $34 (2025) → $37 (2026) → $40 (2027)",
     "Jan 2023 - Dec 2027 (area expansions 2024, 2025)", "85% Towns / 15% City", "Yes - $75 hook-up fee"),
    ("City of Jefferson", "Town of Aztalan", "Paramedic, Advanced, Basic EMS",
     "Per Capita", "$31/capita (2024) → $34 (2025) → $37 (2026) → $40 (2027)",
     "Jan 2024 - Dec 2027", "85% Towns / 15% City", "Yes - $75 hook-up fee"),

    # Fort Atkinson contracts
    ("City of Fort Atkinson", "Town of Jefferson (portions)", "Emergency 911 medical services",
     "Per Capita + CPI", "$7.22/capita base + CPI-W (2-6% annual increase)",
     "Jan 2023 - Dec 2025", "Not specified in contract", "No"),
    ("City of Fort Atkinson", "Town of Koshkonong", "Emergency 911 medical services",
     "Per Capita + CPI", "$7.22/capita base + CPI-W (2-6% annual increase); Pop: 3,524",
     "Jan 2023 - Dec 2025", "Not specified in contract", "No"),

    # Johnson Creek (JCFD) contracts
    ("Johnson Creek (JCFD)", "Town of Aztalan", "Fire + EMS including Ambulance",
     "Equalized Improvement Value", "$32,117 (2024)",
     "Jan 2024 - Dec 2028", "Included in overall budget allocation", "Fire insurance dues remitted to Village"),
    ("Johnson Creek (JCFD)", "Town of Farmington", "Fire + EMS including Ambulance",
     "Equalized Improvement Value", "$118,249 (2024)",
     "Jan 2024 - Dec 2028", "Included in overall budget allocation", "Fire insurance dues remitted"),
    ("Johnson Creek (JCFD)", "Town of Milford", "Fire + EMS including Ambulance",
     "Equalized Improvement Value", "$19,788 (2024)",
     "Jan 2024 - Dec 2028", "Included in overall budget allocation", "Fire insurance dues remitted"),
    ("Johnson Creek (JCFD)", "Town of Watertown", "Fire + EMS including Ambulance",
     "Equalized Improvement Value", "$44,949 (2024)",
     "Jan 2024 - Dec 2028", "Included in overall budget allocation", "Fire insurance dues remitted"),

    # Edgerton FPD
    ("Edgerton FPD", "Town of Koshkonong (portions)", "Fire + EMS",
     "Fixed + Annual Increase", "$10,974 (2024) + greater of: net new construction avg OR CPI+2%",
     "Feb 2023 - auto-renews annually", "EFPD operates and cares for apparatus", "2% fire dues remitted to EFPD"),

    # Lake Mills / Ryan Bros contracts
    ("Lake Mills / Ryan Bros", "Town of Aztalan", "Emergency Ambulance Services",
     "Per Capita + Equalization", "$48/capita (2024); annual increase 3-6% (Ambulance Inflation Factor); equalization credit applied",
     "Effective Jan 1, 2024", "Not specified (Ryan Bros owns equipment)", "No"),
    ("Lake Mills / Ryan Bros", "Town of Oakland", "Emergency Ambulance Services",
     "Per Capita (equal rate across service area)", "Per capita rate equal to Town of Oakland rate (amount TBD)",
     "Effective Jan 1, 2025", "Not specified (Ryan Bros owns equipment)", "No"),
    ("Lake Mills / Ryan Bros", "Town of Lake Mills", "Emergency Ambulance Services",
     "Per Capita (equal rate across service area)", "Per capita rate equal to Town of Lake Mills rate (amount TBD)",
     "Effective Jan 1, 2025", "Not specified (Ryan Bros owns equipment)", "No"),

    # Waterloo
    ("Waterloo Fire & Rescue", "Town of Milford", "Fire + Ambulance (ALS/BLS)",
     "Per Capita + Equalized Value", "$22/capita (2024) → $26/capita (2025)",
     "Jun 2023 - Dec 2025 (auto-renew)", "City retains ownership of all equipment", "Patient billing by Fire Dept"),
    ("Waterloo Fire & Rescue", "Town of Waterloo", "Fire + Ambulance (ALS/BLS)",
     "Per Capita", "$22/capita (2024) → $26/capita (2025)",
     "Jun 2023 - Dec 2025 (auto-renew)", "City retains ownership", "Patient billing by Fire Dept"),

    # Watertown
    ("Watertown Fire Dept", "Town of Milford (portions)", "EMS coverage (1 ambulance, 2 personnel)",
     "Per Capita", "$40/capita (133 pop = $5,320/yr; 6-month contract: $2,660)",
     "Jul 2023 - Dec 2023", "City of Watertown owns equipment", "Watertown bills for transport"),

    # Ixonia
    ("Town of Ixonia Fire & Medical", "Town of Watertown", "Fire + EMS",
     "Formula (Budget x weighted %)", "$49,169 (2025); formula: 1/3 calls + 1/3 population + 1/3 equalized value",
     "Jan 2025 - Dec 2025 (extendable 2yr)", "Excluded from formula (capital costs excluded)", "Yes - Ixonia bills individuals; uncollected → Town pays"),
]

for r, row_data in enumerate(contracts, 2):
    for c, val in enumerate(row_data, 1):
        cell = ws2.cell(row=r, column=c, value=val)
        cell.font = NORMAL
        cell.border = thin_border
        cell.alignment = Alignment(wrap_text=True, vertical='top')

auto_width(ws2)
ws2.column_dimensions['E'].width = 50
ws2.column_dimensions['F'].width = 35

# ============================================================
# TAB 3: PER CAPITA RATE COMPARISON
# ============================================================
ws3 = wb.create_sheet("Per Capita Rates")
ws3.sheet_properties.tabColor = "BF8F00"

headers3 = ["Provider → Town", "Payment Model", "2023 Rate", "2024 Rate", "2025 Rate",
            "2026 Rate", "2027 Rate", "Annual Escalation", "Notes"]
for i, h in enumerate(headers3, 1):
    ws3.cell(row=1, column=i, value=h)
style_header(ws3, 1, 9)

rates = [
    ("", "", "", "", "", "", "", "", ""),
    ("CITY OF JEFFERSON (EMS ONLY)", "", "", "", "", "", "", "", ""),
    ("Jefferson → Town of Jefferson", "Per Capita", "$28.00", "$31.00", "$34.00", "$37.00", "$40.00", "+$3/yr", "Identical contract for all 5 towns"),
    ("Jefferson → Town of Farmington", "Per Capita", "$28.00", "$31.00", "$34.00", "$37.00", "$40.00", "+$3/yr", ""),
    ("Jefferson → Town of Hebron", "Per Capita", "$28.00", "$31.00", "$34.00", "$37.00", "$40.00", "+$3/yr", ""),
    ("Jefferson → Town of Oakland", "Per Capita", "$28.00", "$31.00", "$34.00", "$37.00", "$40.00", "+$3/yr", "Service area expanded 2024 & 2025"),
    ("Jefferson → Town of Aztalan", "Per Capita", "N/A", "$31.00", "$34.00", "$37.00", "$40.00", "+$3/yr", "Started 2024"),
    ("", "", "", "", "", "", "", "", ""),
    ("CITY OF FORT ATKINSON (EMS ONLY)", "", "", "", "", "", "", "", ""),
    ("Fort Atkinson → Town of Jefferson", "Per Capita + CPI", "$7.22 + CPI", "$7.22 + CPI", "$7.22 + CPI", "Contract ends", "", "CPI-W: min 2%, max 6%", "Pop: 318; covers portions only"),
    ("Fort Atkinson → Town of Koshkonong", "Per Capita + CPI", "$7.22 + CPI", "$7.22 + CPI", "$7.22 + CPI", "Contract ends", "", "CPI-W: min 2%, max 6%", "Pop: 3,524"),
    ("", "", "", "", "", "", "", "", ""),
    ("LAKE MILLS / RYAN BROS (EMS ONLY)", "", "", "", "", "", "", "", ""),
    ("Lake Mills → Town of Aztalan", "Per Capita + Equalization", "N/A", "$48.00", "$48 + inflation", "", "", "3-6% (Ambulance Inflation Factor)", "Equalization credit reduces effective rate"),
    ("Lake Mills → Town of Oakland", "Per Capita", "N/A", "N/A", "TBD (equal rate)", "", "", "Equal across service area", "Starts Jan 2025"),
    ("Lake Mills → Town of Lake Mills", "Per Capita", "N/A", "N/A", "TBD (equal rate)", "", "", "Equal across service area", "Starts Jan 2025"),
    ("", "", "", "", "", "", "", "", ""),
    ("WATERLOO (FIRE + EMS)", "", "", "", "", "", "", "", ""),
    ("Waterloo → Town of Milford", "Per Capita", "$18.00", "$22.00", "$26.00", "", "", "+$4/yr", ""),
    ("Waterloo → Town of Waterloo", "Per Capita", "$18.00", "$22.00", "$26.00", "", "", "+$4/yr", ""),
    ("", "", "", "", "", "", "", "", ""),
    ("WATERTOWN (EMS ONLY)", "", "", "", "", "", "", "", ""),
    ("Watertown → Town of Milford", "Per Capita", "$40.00", "", "", "", "", "N/A", "6-month contract; pop 133"),
    ("", "", "", "", "", "", "", "", ""),
    ("IXONIA (FIRE + EMS)", "", "", "", "", "", "", "", ""),
    ("Ixonia → Town of Watertown", "Formula-based", "N/A", "N/A", "$49,169 lump sum", "", "", "1/3 calls + 1/3 pop + 1/3 EV", "Capital costs excluded from formula"),
    ("", "", "", "", "", "", "", "", ""),
    ("JOHNSON CREEK (FIRE + EMS - EQUALIZED VALUE MODEL)", "", "", "", "", "", "", "", ""),
    ("JCFD → Town of Aztalan", "Equalized Value", "$97,636", "$32,117", "", "", "", "Based on EV ratio", "Significant drop 2023→2024"),
    ("JCFD → Town of Farmington", "Equalized Value", "$99,646", "$118,249", "", "", "", "Based on EV ratio", ""),
    ("JCFD → Town of Milford", "Equalized Value", "$16,102", "$19,788", "", "", "", "Based on EV ratio", ""),
    ("JCFD → Town of Watertown", "Equalized Value", "$39,540", "$44,949", "", "", "", "Based on EV ratio", ""),
    ("", "", "", "", "", "", "", "", ""),
    ("EDGERTON FPD (FIRE + EMS)", "", "", "", "", "", "", "", ""),
    ("EFPD → Town of Koshkonong", "Fixed + CPI", "$10,056", "$10,974+", "", "", "", "Greater of: NNC avg or CPI+2%", "Small service area (4 sections)"),
    ("", "", "", "", "", "", "", "", ""),
    ("COMPARISON: COUNTYWIDE SYSTEMS", "", "", "", "", "", "", "", ""),
    ("Portage County (all residents)", "County EMS Levy", "~$55/capita", "~$57/capita", "~$57/capita", "", "", "Levy-funded", "$4,043,061 budget / 70,521 pop"),
    ("Bayfield County (all residents)", "County EMS Levy", "N/A", "~$12/capita", "~$31/capita", "", "", "Exempt from levy lid", "$458k / 15,000 pop (2025 proposed)"),
]

for r, row_data in enumerate(rates, 2):
    for c, val in enumerate(row_data, 1):
        cell = ws3.cell(row=r, column=c, value=val)
        cell.font = NORMAL
        cell.border = thin_border
        cell.alignment = Alignment(wrap_text=True, vertical='top')
    # Style section headers
    if any(row_data[0].startswith(prefix) for prefix in ("CITY OF", "LAKE MILLS", "WATERLOO", "WATERTOWN", "IXONIA", "JOHNSON", "EDGERTON", "COMPARISON")):
        style_section(ws3, r, 9)

auto_width(ws3)

# ============================================================
# TAB 4: PORTAGE COUNTY FINANCIAL DETAIL
# ============================================================
ws4 = wb.create_sheet("Portage County Detail")
ws4.sheet_properties.tabColor = "7030A0"

headers4 = ["Metric", "Value", "Notes"]
for i, h in enumerate(headers4, 1):
    ws4.cell(row=1, column=i, value=h)
style_header(ws4, 1, 3)

portage_data = [
    ("", "", ""),
    ("SYSTEM OVERVIEW", "", ""),
    ("Model", "Hybrid - County contracts with 3 municipal fire departments", "EMS Division within Sheriff's Office"),
    ("Population Served", "~70,521", "Town of Grant contracts separately"),
    ("Area Covered", "~823 square miles", ""),
    ("EMS Coordinator", "Yes - full-time position since 2009", "Part of Sheriff's Office EMS Division"),
    ("Medical Director", "Dr. Michael Clark + 2 Associate Directors", ""),
    ("", "", ""),
    ("CONTRACTED AMBULANCE PROVIDERS (2025)", "", ""),
    ("Stevens Point Fire Dept (SPFD)", "$2,447,541", "2 frontline ambulances 24/7 (2 Paramedics each); 2 reserve ambulances"),
    ("Plover Fire Dept (PFD)", "$1,001,648", "1 frontline ambulance 24/7 (2 Paramedics)"),
    ("Amherst Fire District (AFD)", "$593,872", "1 frontline ambulance 24/7 (min 1 Paramedic + 1 EMT)"),
    ("TOTAL CONTRACTED PERSONNEL", "$4,043,061", "2025 budget for contracted personnel only"),
    ("", "", ""),
    ("AMBULANCE FLEET", "", ""),
    ("Frontline Ambulances", "4", "Staffed 24/7"),
    ("Reserve Ambulances", "2", "Staffed when available (SPFD)"),
    ("Coverage Model", "4 stations cover Portage County", ""),
    ("", "", ""),
    ("REVENUE & BILLING (2024)", "", ""),
    ("Gross Charges", "$5,841,808", ""),
    ("Total Revenue Collected", "$2,050,116", "35% collection rate"),
    ("Medicare (40.38%)", "$855,299 collected on $3,288,008 charges", "Avg $390/trip"),
    ("Private Insurance (8.72%)", "$243,753 collected on $665,704 charges", "Avg $701/trip"),
    ("Private Pay (42.66%)", "$134,867 collected on $1,087,676 charges", "Avg $46/trip"),
    ("Medicaid (6.72%)", "$120,674 collected on $485,916 charges", "Avg $318/trip"),
    ("", "", ""),
    ("CALL VOLUME (2024)", "", ""),
    ("Total Billable Calls", "3,993", "Up from 3,821 in 2023"),
    ("ALS Calls", "2,542", "64% of calls"),
    ("BLS Calls", "1,451", "36% of calls"),
    ("Trend", "Growing: 3,219 (2014) → 3,993 (2024)", "24% increase over 10 years"),
    ("", "", ""),
    ("EMR GROUPS", "", ""),
    ("Number of EMR Groups", "11 municipal-based groups", "Almond, Amherst, Bancroft, Dewey, Grant, Hull, Plover, Rosholt, Rudolph, Sharon, Stockton"),
    ("Training", "~70 hour EMR course via Mid-State Technical College", "County assists with initial training"),
    ("", "", ""),
    ("COUNTY EMS GRANTS", "", ""),
    ("Annual grants to local departments", "Yes", "Amount not specified in report"),
]

for r, row_data in enumerate(portage_data, 2):
    for c, val in enumerate(row_data, 1):
        cell = ws4.cell(row=r, column=c, value=val)
        cell.font = NORMAL
        cell.border = thin_border
        cell.alignment = Alignment(wrap_text=True, vertical='top')
    if row_data[0] in ("SYSTEM OVERVIEW", "CONTRACTED AMBULANCE PROVIDERS (2025)", "AMBULANCE FLEET",
                        "REVENUE & BILLING (2024)", "CALL VOLUME (2024)", "EMR GROUPS", "COUNTY EMS GRANTS"):
        style_section(ws4, r, 3)

auto_width(ws4)

# ============================================================
# TAB 5: BAYFIELD COUNTY FINANCIAL DETAIL
# ============================================================
ws5 = wb.create_sheet("Bayfield County Detail")
ws5.sheet_properties.tabColor = "C00000"

headers5 = ["Metric", "Value", "Notes"]
for i, h in enumerate(headers5, 1):
    ws5.cell(row=1, column=i, value=h)
style_header(ws5, 1, 3)

bayfield_data = [
    ("", "", ""),
    ("SYSTEM OVERVIEW", "", ""),
    ("Model", "County-supported (stipends to 9 providers)", "Established countywide EMS in 2023"),
    ("Population Served", "~15,000", ""),
    ("Area Covered", "2,000+ square miles (land and water)", "Vast rural area - distance is major factor"),
    ("EMS Agencies", "9 long-standing agencies", "Most operate as volunteer responders"),
    ("Staffing Model", "Predominantly volunteer", "Recruitment has struggled for years"),
    ("", "", ""),
    ("2024 BUDGET", "", ""),
    ("Total EMS Allocation", "$185,400", ""),
    ("Purpose", "Support existing agencies + recruitment/awareness campaigns", "Social media campaigns"),
    ("", "", ""),
    ("2025 PROPOSED BUDGET", "", ""),
    ("Total Proposed", "$458,000", "148% increase from 2024"),
    ("   Support for existing agencies", "$185,000", "Stipends to 9 EMS providers"),
    ("   EMS Coordinator", "$145,000", "New position - coordinate schedules, training, compliance"),
    ("   Emergency Medical Dispatch", "$128,000", "Newly established in 2025"),
    ("Cost to Homeowner", "$21.46/year per $200,000 home", "Mill rate: 0.1073 mills"),
    ("Levy Status", "Exempt from levy lid", "Wis. Stat. § 66.0602(3)(e)6"),
    ("", "", ""),
    ("PHASE 1 PROPOSAL (Strategic Management & Consulting)", "", ""),
    ("Proposed Model", "6-person countywide paramedic response crew", "Non-transporting agency"),
    ("Vehicles", "Response vehicles (NOT ambulances)", "Stationed around county"),
    ("Use Cases", "Supplement scenes, ALS when needed, backup for single responders", ""),
    ("Coordinator Role", "Interact with agencies, coordinate training, grant-finding, compliance", ""),
    ("", "", ""),
    ("KEY CONTEXT", "", ""),
    ("Background", "EMS Study Committee formed in 2021", "Final report: bayfieldcounty.wi.gov"),
    ("Study", "Draft released mid-August 2024", "Strategic Management & Consulting, LLC"),
    ("Core Issue", "Volunteer recruitment shortfalls + increased mutual aid requests", "Goal: speedy response times"),
]

for r, row_data in enumerate(bayfield_data, 2):
    for c, val in enumerate(row_data, 1):
        cell = ws5.cell(row=r, column=c, value=val)
        cell.font = NORMAL
        cell.border = thin_border
        cell.alignment = Alignment(wrap_text=True, vertical='top')
    if row_data[0] in ("SYSTEM OVERVIEW", "2024 BUDGET", "2025 PROPOSED BUDGET",
                        "PHASE 1 PROPOSAL (Strategic Management & Consulting)", "KEY CONTEXT"):
        style_section(ws5, r, 3)

auto_width(ws5)

# ============================================================
# TAB 6: AMBULANCE FLEET COMPARISON
# ============================================================
ws6 = wb.create_sheet("Ambulance Fleet")
ws6.sheet_properties.tabColor = "ED7D31"

headers6 = ["Department", "Unit ID", "Year", "Make/Manufacturer", "Chassis", "EMS Level", "AEDs", "O2 Bottles",
            "Backboards", "Stair Chairs", "SCBA", "Staffing", "Age (yrs)", "Notes"]
for i, h in enumerate(headers6, 1):
    ws6.cell(row=1, column=i, value=h)
style_header(ws6, 1, 14)

# Ambulance data extracted from MABAS sheets
ambulance_data = [
    ("", "", "", "", "", "", "", "", "", "", "", "", "", ""),
    ("JEFFERSON COUNTY DEPARTMENTS", "", "", "", "", "", "", "", "", "", "", "", "", ""),
    ("Fort Atkinson", "Medic 8158", 2023, "Lifeline", "Ford", "ALS - Paramedic", 1, 3, 3, 1, "MSA (2x, 4500 PSI)", "1 EMT-P + 1 AEMT", 3, "1st out ambulance; newest in county"),
    ("Fort Atkinson", "Rescue 8159", 2017, "LSV", "Ford", "ALS - Paramedic", 1, 3, 3, 1, "N/A", "1 EMT-P + 1 AEMT", 9, "2nd out ambulance"),
    ("Fort Atkinson", "Rescue 8157", 2004, "EDM", "Chevrolet", "BLS", 1, 3, 3, 1, "N/A", "BLS crew", 22, "3rd out; 22 YEARS OLD - replacement priority"),
    ("", "", "", "", "", "", "", "", "", "", "", "", "", ""),
    ("Jefferson", "Rescue 754", 2021, "Horton", "Ford", "ALS - Paramedic", 1, 3, 3, 1, "N/A", "1 EMT + 1 AEMT + 1 Driver", 5, "1st out ambulance"),
    ("Jefferson", "Rescue 755", 2014, "Horton", "Ford", "ALS - Paramedic", 1, 3, 3, 1, "N/A", "1 EMT + 1 AEMT + 1 Driver", 12, "2nd out"),
    ("Jefferson", "Rescue 756", 2007, "Horton", "Ford E-350", "Paramedic", 1, 5, 3, 1, "N/A", "2 Paramedics", 19, "MABAS unit; 19 years old"),
    ("Jefferson", "Rescue 799", 2019, "Interceptor", "Ford", "Paramedic Intercept", 0, 0, 0, 0, "N/A", "Paramedic", 7, "Paramedic intercept vehicle (not ambulance)"),
    ("Jefferson", "Rescue 798", None, "Unknown", "Unknown", "Unknown", 0, 0, 0, 0, "N/A", "Unknown", None, "Limited data available"),
    ("", "", "", "", "", "", "", "", "", "", "", "", "", ""),
    ("Watertown", "Med 54 / 4154", 2023, "Lifeline", "F350", "Paramedic", 1, 5, 0, 0, "MSA (2x, 2216 PSI)", "2 Paramedics", 3, "1st out"),
    ("Watertown", "Med 52 / 4152", 2006, "International", "4300", "Paramedic", 1, 5, 0, 0, "MSA (2x, 2216 PSI)", "2 Paramedics", 20, "20 YEARS OLD - replacement priority"),
    ("Watertown", "Med 53 / 4153", 2014, "Ford", "F450", "Paramedic", 1, 5, 0, 0, "MSA (2x, 2216 PSI)", "2 Paramedics", 12, "3rd out"),
    ("", "", "", "", "", "", "", "", "", "", "", "", "", ""),
    ("Waterloo Fire & Rescue", "Rescue 3959", 2005, "Freightliner / MedTech", "Freightliner", "Advanced EMT", 0, 0, 4, 1, "N/A", "1 Driver + 1 Crew Chief + 1 EMT", 21, "21 YEARS OLD - replacement priority"),
    ("Waterloo Fire & Rescue", "Rescue 3958", 2014, "Freightliner / Horton", "Freightliner", "Advanced EMT", 0, 0, 4, 1, "N/A", "1 Driver + 1 Crew Chief + 1 EMT", 12, ""),
    ("", "", "", "", "", "", "", "", "", "", "", "", "", ""),
    ("Ixonia", "8351", 2012, "Lifeline", "Ford F-550", "EMT / EMT-I", 1, 3, 3, 1, "N/A", "1 EMT + 1 EMT-I", 14, "Single ambulance for department"),
    ("", "", "", "", "", "", "", "", "", "", "", "", "", ""),
    ("Palmyra", "Rescue 717", None, "Unknown", "Unknown", "EMT / AEMT", 0, 0, 0, 0, "MSA G1 (2x, 4500 PSI)", "EMT + AEMT", None, "Minimal data in MABAS sheet"),
    ("", "", "", "", "", "", "", "", "", "", "", "", "", ""),
    ("Cambridge", "None listed", None, "", "", "", "", "", "", "", "", "", None, "NO ambulance/EMS tab in MABAS sheet"),
    ("Helenville", "None listed", None, "", "", "", "", "", "", "", "", "", None, "NO ambulance/EMS tab; fire only"),
    ("Johnson Creek", "None listed", None, "", "", "", "", "", "", "", "", "", None, "NO ambulance/EMS tab; fire only"),
    ("Lake Mills", "None listed", None, "", "", "", "", "", "", "", "", "", None, "EMS via Ryan Bros (private contractor)"),
    # Rome and Sullivan are fire-only — not EMS providers, excluded from EMS comparison
    ("Western Lakes", "Template only", None, "", "", "", "", "", "", "", "", "", None, "4 ambulance slots in sheet but ALL EMPTY"),
    ("Whitewater", "3 ambulances", None, "", "", "ALS - Paramedic (since 2023)", "", "", "", "", "", "18 FT + 12 POC; 5 per 24hr shift", None, "MABAS sheet empty but city website confirms 3 ambulances; 2,000+ calls/yr; 100 sq mi service area"),
    ("", "", "", "", "", "", "", "", "", "", "", "", "", ""),
    ("PORTAGE COUNTY (COMPARISON)", "", "", "", "", "", "", "", "", "", "", "", "", ""),
    ("Stevens Point FD (SPFD)", "Med 2", None, "", "", "Paramedic", "", "", "", "", "", "2 Paramedics", None, "4 ambulances total (Med 2-5); 2 frontline 24/7 + 2 reserve"),
    ("Stevens Point FD (SPFD)", "Med 3", None, "", "", "Paramedic", "", "", "", "", "", "2 Paramedics", None, "Also: 2 engines (2018/2025 Pierce), 2 trucks, 1 tower"),
    ("Stevens Point FD (SPFD)", "Med 4", None, "", "", "Paramedic", "", "", "", "", "", "2 Paramedics", None, "Reserve unit"),
    ("Stevens Point FD (SPFD)", "Med 5", None, "", "", "Paramedic", "", "", "", "", "", "2 Paramedics", None, "Reserve unit"),
    ("Plover FD (PFD)", "54 Med 1", None, "", "", "Paramedic", "", "", "", "", "", "2 Paramedics", None, "1 ambulance 24/7; Also: 2 engines, 1 platform, 2 tenders, 1 brush"),
    ("Amherst Fire District (AFD)", "2 ambulances", None, "", "", "Paramedic", "", "", "", "", "", "Min 1 Paramedic + 1 EMT", None, "1 ambulance 24/7; $593,872 contract; detail not available online"),
    ("", "", "", "", "", "", "", "", "", "", "", "", "", ""),
    ("BAYFIELD COUNTY (COMPARISON)", "", "", "", "", "", "", "", "", "", "", "", "", ""),
    ("9 agencies countywide", "Not specified", None, "", "", "Varies", "", "", "", "", "", "Mostly volunteers", None, "Proposed: 6-person paramedic response crew with response vehicles"),
]

for r, row_data in enumerate(ambulance_data, 2):
    for c, val in enumerate(row_data, 1):
        cell = ws6.cell(row=r, column=c, value=val)
        cell.font = NORMAL
        cell.border = thin_border
        cell.alignment = Alignment(wrap_text=True, vertical='top')
    if row_data[0] in ("JEFFERSON COUNTY DEPARTMENTS", "PORTAGE COUNTY (COMPARISON)", "BAYFIELD COUNTY (COMPARISON)"):
        style_section(ws6, r, 14)

auto_width(ws6)
ws6.column_dimensions['N'].width = 45

# ============================================================
# TAB 7: FIRE APPARATUS INVENTORY
# ============================================================
ws7b = wb.create_sheet("Fire Apparatus")
ws7b.sheet_properties.tabColor = "FF6600"

headers7b = ["Department", "Data Year", "Engines", "Trucks/Ladders", "Squads", "Tenders",
             "Command", "Brush", "ATVs", "Watercraft", "Trailers/Utility", "Ambulances",
             "Total Units", "Has EMS Tab"]
for i, h in enumerate(headers7b, 1):
    ws7b.cell(row=1, column=i, value=h)
style_header(ws7b, 1, 14)

fire_data = [
    ("Cambridge", "~2024", 3, 0, 0, 1, 0, 1, 2, 1, 0, 0, 8, "No"),
    ("Fort Atkinson", "2020", 3, 1, 2, 2, 1, 1, 1, 2, 0, 3, 16, "Yes"),
    ("Helenville", "2016", 2, 0, 1, 2, 0, 1, 1, 0, 0, 0, 7, "No"),
    ("Ixonia", "2020", 2, 0, 0, 2, 0, 1, 1, 0, 0, 1, 7, "Yes"),
    ("Jefferson", "2024", 3, 1, 1, 2, 1, 1, 1, 2, 3, 5, 20, "Yes"),
    ("Johnson Creek", "2016", 2, 1, 1, 2, 1, 1, 1, 0, 0, 0, 9, "No"),
    ("Lake Mills", "2020", 3, 1, 1, 1, 1, 1, 0, 2, 1, 0, 11, "No"),
    ("Palmyra", "2025", 1, 1, 0, 1, 1, 0, 1, 0, 1, 1, 7, "Yes"),
    # Rome and Sullivan are fire-only departments — excluded from EMS comparison
    ("Waterloo Fire & Rescue", "2015", 2, 1, 0, 2, 0, 1, 2, 0, 0, 2, 10, "Yes"),
    ("Watertown", "~2020", 3, 1, 0, 2, 1, 1, 0, 1, 1, 3, 12, "Yes (3 ambulances)"),
    ("Western Lakes", "~2024", 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, "Template only - NO DATA"),
    ("Whitewater", "~2024", 3, 1, 2, 2, 1, 1, 1, 0, 1, 3, 15, "MABAS empty; city website: 3 ambulances + 10 fire apparatus"),
    ("", "", "", "", "", "", "", "", "", "", "", "", "", ""),
    ("JEFFERSON CO TOTAL", "", 31, 8, 8, 22, 7, 12, 13, 8, 7, 18, "~145", ""),
    ("", "", "", "", "", "", "", "", "", "", "", "", "", ""),
    ("Portage County", "2024-2025", "4 (SPFD 2, PFD 2)", "3 (SPFD 2 trucks + 1 tower, PFD 1 platform)", "N/A", "2 (PFD)", "4+ (SPFD)", "1 (PFD)", "N/A", "N/A", "N/A", 7, "~35+ est.", "SPFD + PFD + AFD"),
    ("Bayfield County", "2025", "N/A", "N/A", "N/A", "N/A", "N/A", "N/A", "N/A", "N/A", "N/A", "TBD", "TBD", "9 volunteer agencies"),
]

for r, row_data in enumerate(fire_data, 2):
    for c, val in enumerate(row_data, 1):
        cell = ws7b.cell(row=r, column=c, value=val)
        cell.font = NORMAL
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', wrap_text=True, vertical='top')
    ws7b.cell(row=r, column=1).alignment = Alignment(wrap_text=True, vertical='top')

# Highlight the totals row
for c in range(1, 15):
    ws7b.cell(row=17, column=c).font = BOLD

auto_width(ws7b)

# ============================================================
# TAB 7: LEVY LIMIT ANALYSIS
# ============================================================
ws7 = wb.create_sheet("Levy Limit Analysis")
ws7.sheet_properties.tabColor = "FF0000"

headers7 = ["Topic", "Details", "Implication for Jefferson County"]
for i, h in enumerate(headers7, 1):
    ws7.cell(row=1, column=i, value=h)
style_header(ws7, 1, 3)

levy_data = [
    ("", "", ""),
    ("CURRENT STATE", "", ""),
    ("Jefferson County levy role", "County has NO role in EMS; each municipality funds its own EMS from its own levy", "Each municipality independently at/near levy cap for EMS spending"),
    ("Municipal levy caps", "Cities, villages, and towns each have individual levy limits they cannot exceed", "EMS costs compete with other municipal services for limited levy room"),
    ("", "", ""),
    ("COUNTYWIDE EMS LEVY EXCEPTION", "", ""),
    ("Wis. Stat. § 66.0602(3)(e)6", "The levy limit does NOT apply to 'the amount that a county levies in that year for a countywide emergency medical system'", "If Jefferson County adopts countywide EMS, the county levy for EMS is EXEMPT from levy caps"),
    ("County advantage", "County has more levy leverage than individual municipalities", "Could shift EMS funding burden from municipalities to county without hitting municipal levy caps"),
    ("", "", ""),
    ("JOINT FIRE/EMS EXCEPTION", "", ""),
    ("Wis. Stat. § 66.0602(3)(h)", "Municipalities in joint fire/EMS districts can adjust their levy limit if assessed charges increase ≤ CPI + 2%", "Existing joint arrangements (like JCFD) may already qualify for this adjustment"),
    ("First year limitation", "Cannot use exception in the first year of formation (no prior year comparison)", "New joint districts would not benefit until year 2"),
    ("Multiple districts", "A municipality can use the exception for BOTH joint fire AND joint EMS if they are separate districts", "Potential for stacking adjustments"),
    ("", "", ""),
    ("CASE STUDY: LAFAYETTE COUNTY", "", ""),
    ("Their approach", "$5.00/capita countywide levy distributed to municipalities based on population", "~$87,000 total fiscal impact (much smaller county)"),
    ("Controversy", "Blanchardville + 12 subdivisions filed petition challenging the levy exception", "Risk: legal challenges possible when implementing"),
    ("Resolution", "Agreement reached, EMS Advisory Subcommittee created; 5-year term with option to extend", "Governance structure model for Jefferson County"),
    ("", "", ""),
    ("CASE STUDY: PORTAGE COUNTY", "", ""),
    ("Their approach", "County EMS levy funds $4M+ in contracted paramedic services from 3 municipal fire depts", "Much larger financial commitment, covers 70,521 people"),
    ("Structure", "EMS Division within Sheriff's Office, full-time EMS Coordinator", "Administrative infrastructure in place"),
    ("", "", ""),
    ("CASE STUDY: BAYFIELD COUNTY", "", ""),
    ("Their approach", "Countywide EMS established 2023, levy exempt from lid; $458k (2025)", "Smaller budget, focused on stabilizing volunteer agencies"),
    ("Progression", "$185,400 (2024) → $458,000 (2025) = 148% increase in one year", "Shows rapid scaling possible once framework established"),
]

for r, row_data in enumerate(levy_data, 2):
    for c, val in enumerate(row_data, 1):
        cell = ws7.cell(row=r, column=c, value=val)
        cell.font = NORMAL
        cell.border = thin_border
        cell.alignment = Alignment(wrap_text=True, vertical='top')
    if row_data[0] in ("CURRENT STATE", "COUNTYWIDE EMS LEVY EXCEPTION", "JOINT FIRE/EMS EXCEPTION",
                        "CASE STUDY: LAFAYETTE COUNTY", "CASE STUDY: PORTAGE COUNTY", "CASE STUDY: BAYFIELD COUNTY"):
        style_section(ws7, r, 3)

auto_width(ws7)

# Save
output_path = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                           "Jefferson_Co_EMS_Financial_Comparison.xlsx")
wb.save(output_path)
print(f"Spreadsheet saved to: {output_path}")
